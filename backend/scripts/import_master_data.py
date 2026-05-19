#!/usr/bin/env python3
"""
Import master data from Excel into the ERP database.

Usage:
    python -m scripts.import_master_data
    python -m scripts.import_master_data --db-name bhspl_scm_uat
    python -m scripts.import_master_data --excel "C:/path/to/file.xlsx"
"""

import argparse
import sys
import os
from datetime import datetime
from decimal import Decimal, InvalidOperation
from urllib.parse import quote_plus

import openpyxl
from sqlalchemy import create_engine, text


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_bool(val) -> bool:
    """Y / y / Yes / 1 / TRUE -> True, everything else -> False."""
    if val is None:
        return False
    return str(val).strip().lower() in ("y", "yes", "1", "true")


def parse_decimal(val, default=Decimal("0")) -> Decimal:
    """Convert to Decimal; treat empty / NA / N / None as *default*."""
    if val is None:
        return default
    s = str(val).strip()
    if s == "" or s.upper() in ("NA", "N", "-", "N/A", "NONE"):
        return default
    try:
        return Decimal(s)
    except InvalidOperation:
        return default


def parse_int(val, default=0) -> int:
    d = parse_decimal(val, Decimal(str(default)))
    return int(d)


def clean_str(val) -> str | None:
    """Return stripped string or None."""
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s.upper() in ("NA", "N/A", "NONE", "-"):
        return None
    return s


def parse_date(val):
    """Return a datetime or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if s == "" or s.upper() in ("NA", "N/A", "NONE", "-"):
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Item-type mapping
# ---------------------------------------------------------------------------

ITEM_TYPE_MAP = {
    "tablets & capsules": "medicine",
    "syrups & suspensions": "medicine",
    "injectables": "medicine",
    "other medicines": "medicine",
    "vet-injectables": "medicine",
    "bolus & boli": "medicine",
    "vet-medicines": "medicine",
    "medical equipment": "asset",
    "other equipment": "asset",
    "consumables": "consumable",
    "surgical": "consumable",
}


def map_item_type(raw: str | None) -> str:
    if raw is None:
        return "traded"
    return ITEM_TYPE_MAP.get(raw.strip().lower(), "traded")


# ---------------------------------------------------------------------------
# Barcode-type mapping
# ---------------------------------------------------------------------------

BARCODE_TYPE_VALID = {"qrcode", "barcode_128", "barcode_ean13", "auto"}


def map_barcode_type(raw: str | None) -> str:
    if raw is None:
        return "auto"
    v = raw.strip().lower().replace(" ", "_").replace("-", "_")
    return v if v in BARCODE_TYPE_VALID else "auto"


# ---------------------------------------------------------------------------
# Read one sheet into a list[dict]
# ---------------------------------------------------------------------------

def _normalize_header(h: str) -> str:
    """Strip whitespace, asterisks, and trailing special chars from header."""
    return h.strip().rstrip("*").strip()


def read_sheet(wb: openpyxl.Workbook, sheet_name: str) -> list[dict]:
    """Return rows as list of dicts keyed by normalized header (first row)."""
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [_normalize_header(str(h)) if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        # skip entirely blank rows
        if all(c is None for c in row):
            continue
        result.append(dict(zip(headers, row)))
    return result


# ---------------------------------------------------------------------------
# DB credentials from .env
# ---------------------------------------------------------------------------

def load_env(env_path: str) -> dict:
    """Minimal .env loader (no dependency on python-dotenv)."""
    env = {}
    if not os.path.isfile(env_path):
        return env
    with open(env_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" not in line:
                continue
            key, val = line.split("=", 1)
            env[key.strip()] = val.strip()
    return env


# ---------------------------------------------------------------------------
# Main import logic
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Import master data from Excel")
    parser.add_argument("--db-name", default="bhspl_scm_uat", help="Database name (default: bhspl_scm_uat)")
    parser.add_argument(
        "--excel",
        default=r"C:\Users\saima\Downloads\Master Data - Final.xlsx",
        help="Path to the Excel file",
    )
    args = parser.parse_args()

    # ---- load env ----
    backend_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    env_path = os.path.join(backend_dir, ".env")
    env = load_env(env_path)

    db_host = env.get("DB_HOST", "localhost")
    db_port = env.get("DB_PORT", "3306")
    db_user = env.get("DB_USER", "root")
    db_password = env.get("DB_PASSWORD", "")
    db_name = args.db_name

    db_url = (
        f"mysql+pymysql://{quote_plus(db_user)}:{quote_plus(db_password)}"
        f"@{db_host}:{db_port}/{db_name}?charset=utf8mb4"
    )

    print(f"[*] Connecting to {db_host}:{db_port}/{db_name} ...")
    engine = create_engine(db_url, echo=False)

    # ---- open excel ----
    print(f"[*] Reading Excel: {args.excel}")
    wb = openpyxl.load_workbook(args.excel, data_only=True)
    print(f"    Sheets found: {wb.sheetnames}")

    with engine.begin() as conn:
        # ==================================================================
        # STEP 1 -- FLUSH existing data (FK-safe order)
        # ==================================================================
        print("\n[1] Flushing existing data ...")

        flush_tables = [
            # consumption
            "consumption_items",
            "consumption_entries",
            # stock
            "stock_ledger",
            "stock_balance",
            # grn
            "putaway_items",
            "putaway_orders",
            "quality_inspection_items",
            "quality_inspections",
            "grn_items",
            "goods_receipt_notes",
            # purchase orders
            "purchase_order_items",
            "purchase_orders",
            # material requests
            "material_request_items",
            "material_requests",
            # quotations
            "quotation_items",
            "quotations",
            # indents
            "indent_acknowledgement_items",
            "indent_acknowledgements",
            "indent_items",
            "indents",
            # returns
            "purchase_return_items",
            "purchase_returns",
            # item packing + items
            "price_list_items",
            "item_packing",
            "items",
            # categories
            "item_categories",
            # uom
            "uom_conversions",
            "uom",
            # vendors
            "vendor_items",
            "vendor_contracts",
            "vendor_ratings",
            "vendors",
            # projects (keep organizations)
            "projects",
        ]

        conn.execute(text("SET FOREIGN_KEY_CHECKS = 0"))
        for tbl in flush_tables:
            try:
                conn.execute(text(f"DELETE FROM `{tbl}`"))
                conn.execute(text(f"ALTER TABLE `{tbl}` AUTO_INCREMENT = 1"))
                print(f"    Flushed {tbl}")
            except Exception as e:
                print(f"    WARN: could not flush {tbl}: {e}")
        conn.execute(text("SET FOREIGN_KEY_CHECKS = 1"))
        print("    Done flushing.")

        # ==================================================================
        # STEP 2a -- UOMs
        # ==================================================================
        print("\n[2a] Importing UOMs ...")

        uom_sheet = read_sheet(wb, "UOM")
        uom_name_to_id: dict[str, int] = {}  # lowercase name -> id
        uom_abbr_to_id: dict[str, int] = {}  # lowercase abbr -> id

        for row in uom_sheet:
            name = clean_str(row.get("Name"))
            abbr = clean_str(row.get("Abbreviation"))
            if not name or not abbr:
                continue
            # dedup by name
            if name.lower() in uom_name_to_id:
                continue

            conn.execute(
                text(
                    "INSERT INTO uom (name, abbreviation, is_active, created_at) "
                    "VALUES (:name, :abbr, 1, NOW())"
                ),
                {"name": name, "abbr": abbr},
            )
            uid = conn.execute(text("SELECT LAST_INSERT_ID()")).scalar()
            uom_name_to_id[name.lower()] = uid
            uom_abbr_to_id[abbr.lower()] = uid
            print(f"    UOM: {name} ({abbr}) -> id={uid}")

        print(f"    Imported {len(uom_name_to_id)} UOMs.")

        # ==================================================================
        # STEP 2b -- UOM Conversions (cols E, F, G of same sheet)
        # ==================================================================
        print("\n[2b] Importing UOM Conversions ...")

        conv_count = 0
        for row in uom_sheet:
            from_uom_name = clean_str(row.get("From UOM"))
            to_uom_name = clean_str(row.get("To UOM"))
            factor = parse_decimal(row.get("Conversion Factor"))
            if not from_uom_name or not to_uom_name or factor == 0:
                continue

            from_id = uom_name_to_id.get(from_uom_name.lower()) or uom_abbr_to_id.get(from_uom_name.lower())
            to_id = uom_name_to_id.get(to_uom_name.lower()) or uom_abbr_to_id.get(to_uom_name.lower())

            if not from_id:
                print(f"    WARN: From UOM '{from_uom_name}' not found, skipping conversion.")
                continue
            if not to_id:
                print(f"    WARN: To UOM '{to_uom_name}' not found, skipping conversion.")
                continue

            conn.execute(
                text(
                    "INSERT INTO uom_conversions (from_uom_id, to_uom_id, conversion_factor) "
                    "VALUES (:from_id, :to_id, :factor)"
                ),
                {"from_id": from_id, "to_id": to_id, "factor": float(factor)},
            )
            conv_count += 1
            print(f"    Conversion: {from_uom_name} -> {to_uom_name} x {factor}")

        print(f"    Imported {conv_count} UOM conversions.")

        # ==================================================================
        # STEP 2c -- Projects
        # ==================================================================
        print("\n[2c] Importing Projects ...")

        proj_sheet = read_sheet(wb, "Projects")
        proj_count = 0
        for row in proj_sheet:
            proj_name = clean_str(row.get("Project Name"))
            proj_code = clean_str(row.get("Project Code"))
            if not proj_name or not proj_code:
                continue

            spoc = clean_str(row.get("SPOC"))
            start_date = parse_date(row.get("Start Date"))
            end_date = parse_date(row.get("End Date"))
            status_raw = clean_str(row.get("Status"))
            status = "active"
            if status_raw:
                sl = status_raw.lower()
                if sl in ("active", "inactive", "completed"):
                    status = sl

            conn.execute(
                text(
                    "INSERT INTO projects (organization_id, name, code, description, "
                    "start_date, end_date, status, created_at, updated_at) "
                    "VALUES (:org_id, :name, :code, :desc, :start, :end, :status, NOW(), NOW())"
                ),
                {
                    "org_id": 1,
                    "name": proj_name,
                    "code": proj_code,
                    "desc": spoc,  # store SPOC in description as there is no spoc column
                    "start": start_date,
                    "end": end_date,
                    "status": status,
                },
            )
            proj_count += 1
            print(f"    Project: {proj_code} - {proj_name}")

        print(f"    Imported {proj_count} projects.")

        # ==================================================================
        # STEP 2d -- Item Categories
        # ==================================================================
        print("\n[2d] Importing Item Categories ...")

        cat_sheet = read_sheet(wb, "Item Categories")

        # First pass: insert categories without parent (or all as parent_id=NULL)
        # then second pass: update parent_id via parent_code
        cat_code_to_id: dict[str, int] = {}
        cat_rows_with_parent: list[tuple[str, str]] = []  # (code, parent_code)

        for row in cat_sheet:
            code = clean_str(row.get("Category Code"))
            if not code:
                continue
            item_type = clean_str(row.get("Item Type"))
            parent_code = clean_str(row.get("Parent Code"))
            item_code_col = clean_str(row.get("Item Code"))  # sometimes used as name
            description = clean_str(row.get("Description"))

            # Use description as name; fall back to code
            name = description or item_type or code

            if code.upper() in cat_code_to_id:
                continue  # skip duplicate codes
            conn.execute(
                text(
                    "INSERT IGNORE INTO item_categories (parent_id, name, code, description, level, is_active, created_at) "
                    "VALUES (NULL, :name, :code, :desc, 0, 1, NOW())"
                ),
                {"name": name, "code": code, "desc": description},
            )
            cid = conn.execute(text("SELECT LAST_INSERT_ID()")).scalar()
            if cid:
                cat_code_to_id[code.upper()] = cid

            if parent_code:
                cat_rows_with_parent.append((code, parent_code))

            print(f"    Category: {code} - {name}")

        # Second pass: set parent_id and compute level
        for code, parent_code in cat_rows_with_parent:
            pid = cat_code_to_id.get(parent_code.upper())
            if pid:
                cid = cat_code_to_id[code.upper()]
                # Calculate level by walking up the parent chain
                level = 1
                walk_code = parent_code.upper()
                visited = set()
                while walk_code in cat_code_to_id and walk_code not in visited:
                    visited.add(walk_code)
                    # check if this parent has a parent
                    parent_result = conn.execute(
                        text("SELECT parent_id FROM item_categories WHERE id = :id"),
                        {"id": cat_code_to_id[walk_code]},
                    ).fetchone()
                    if parent_result and parent_result[0]:
                        level += 1
                        # find the code for that parent
                        walk_id = parent_result[0]
                        walk_code_found = None
                        for c, i in cat_code_to_id.items():
                            if i == walk_id:
                                walk_code_found = c
                                break
                        if walk_code_found:
                            walk_code = walk_code_found
                        else:
                            break
                    else:
                        break

                conn.execute(
                    text("UPDATE item_categories SET parent_id = :pid, level = :level WHERE id = :cid"),
                    {"pid": pid, "cid": cid, "level": level},
                )
                print(f"    Linked {code} -> parent {parent_code} (level={level})")
            else:
                print(f"    WARN: Parent code '{parent_code}' not found for category '{code}'")

        print(f"    Imported {len(cat_code_to_id)} categories.")

        # ==================================================================
        # Helper: resolve UOM from text
        # ==================================================================

        def resolve_uom(uom_text: str | None) -> int | None:
            """Try to match UOM by name, abbreviation, or partial match."""
            if not uom_text:
                return None
            key = uom_text.strip().lower()
            # exact name match
            if key in uom_name_to_id:
                return uom_name_to_id[key]
            # exact abbreviation match
            if key in uom_abbr_to_id:
                return uom_abbr_to_id[key]
            # partial match: check if any UOM name contains the text or vice versa
            for name, uid in uom_name_to_id.items():
                if key in name or name in key:
                    return uid
            for abbr, uid in uom_abbr_to_id.items():
                if key in abbr or abbr in key:
                    return uid
            return None

        # ==================================================================
        # STEP 2e -- Items from 4 sheets (deduplicate by item_code)
        # ==================================================================
        print("\n[2e] Importing Items ...")

        # Map sheet names to short prefixes for globally unique item codes
        item_sheets_with_prefix = {
            "APGENCO-SNCU": "SNCU",
            "104": "104",
            "108": "108",
            "1962": "1962",
        }
        seen_item_codes: set[str] = set()
        item_code_to_id: dict[str, int] = {}
        # Collect packing data while iterating items
        packing_data: list[dict] = []  # list of dicts with item_code, packing_name, packing_uom, qty_per_pack
        item_count = 0
        skipped = 0

        for sheet_name, prefix in item_sheets_with_prefix.items():
            if sheet_name not in wb.sheetnames:
                print(f"    WARN: Sheet '{sheet_name}' not found, skipping.")
                continue

            print(f"\n    --- Sheet: {sheet_name} (prefix: {prefix}) ---")
            items_sheet = read_sheet(wb, sheet_name)

            for row in items_sheet:
                raw_code = clean_str(row.get("Item Code"))
                if not raw_code:
                    continue

                # Prefix with project to make globally unique
                item_code = f"{prefix}-{raw_code}"

                # Deduplicate (within same prefixed scope)
                if item_code.upper() in seen_item_codes:
                    skipped += 1
                    continue
                seen_item_codes.add(item_code.upper())

                item_name = clean_str(row.get("Item Name")) or item_code
                description = clean_str(row.get("Description (Composition)"))
                item_type_raw = clean_str(row.get("Item Type"))
                item_type = map_item_type(item_type_raw)
                category_code = clean_str(row.get("Category Code"))
                manufacturer = clean_str(row.get("Manufactures"))
                brand = clean_str(row.get("Brand"))
                primary_uom_text = clean_str(row.get("Primary UOM"))
                secondary_uom_text = clean_str(row.get("Secondary UOM"))
                hsn_code = clean_str(row.get("HSN Code"))
                sku = clean_str(row.get("SKU"))
                barcode_type = map_barcode_type(clean_str(row.get("Barcode Type")))

                has_batch = parse_bool(row.get("Has Batch"))
                has_serial = parse_bool(row.get("Has Serial"))
                has_expiry = parse_bool(row.get("Has Expiry"))

                shelf_life_days = parse_int(row.get("Shelf Life Days"))
                safety_stock = parse_decimal(row.get("Safety Stock"))
                reorder_level = parse_decimal(row.get("Reorder Level"))
                reorder_qty = parse_decimal(row.get("Reorder Qty"))
                lead_time_days = parse_int(row.get("Lead Time Days"))
                min_order_qty = parse_decimal(row.get("Min Order Qty"))
                max_order_qty = parse_decimal(row.get("Max Order Qty"))

                weight = parse_decimal(row.get("Weight"), Decimal("0"))
                weight_uom = clean_str(row.get("Weight UOM"))
                volume = parse_decimal(row.get("Volume"), Decimal("0"))
                volume_uom = clean_str(row.get("Volume UOM"))

                purchase_price = parse_decimal(row.get("Purchase Price"))
                selling_price = parse_decimal(row.get("Selling Price"))
                mrp = parse_decimal(row.get("MRP"))
                tax_rate = parse_decimal(row.get("Tax Rate %"))
                cgst = parse_decimal(row.get("CGST %"))
                sgst = parse_decimal(row.get("SGST %"))
                igst = parse_decimal(row.get("IGST %"))

                # Resolve FK references
                category_id = None
                if category_code:
                    category_id = cat_code_to_id.get(category_code.upper())
                    if not category_id:
                        print(f"    WARN: Category '{category_code}' not found for item '{item_code}'")

                primary_uom_id = resolve_uom(primary_uom_text)
                if not primary_uom_id:
                    # Primary UOM is required - create a fallback UOM
                    fallback_name = primary_uom_text or "Each"
                    fallback_abbr = (primary_uom_text or "Ea")[:10]
                    print(f"    WARN: UOM '{primary_uom_text}' not found for item '{item_code}', creating '{fallback_name}'")
                    conn.execute(
                        text(
                            "INSERT INTO uom (name, abbreviation, is_active, created_at) "
                            "VALUES (:name, :abbr, 1, NOW())"
                        ),
                        {"name": fallback_name, "abbr": fallback_abbr},
                    )
                    primary_uom_id = conn.execute(text("SELECT LAST_INSERT_ID()")).scalar()
                    uom_name_to_id[fallback_name.lower()] = primary_uom_id
                    uom_abbr_to_id[fallback_abbr.lower()] = primary_uom_id

                secondary_uom_id = resolve_uom(secondary_uom_text)

                conn.execute(
                    text(
                        "INSERT IGNORE INTO items ("
                        "  category_id, item_code, name, description, item_type, "
                        "  primary_uom_id, secondary_uom_id, hsn_code, sku, barcode_type, "
                        "  has_batch, has_serial, has_expiry, shelf_life_days, "
                        "  safety_stock, reorder_level, reorder_qty, lead_time_days, "
                        "  min_order_qty, max_order_qty, weight, weight_uom, volume, volume_uom, "
                        "  purchase_price, selling_price, mrp, tax_rate, cgst_rate, sgst_rate, igst_rate, "
                        "  brand, manufacturer, is_active, created_at, updated_at"
                        ") VALUES ("
                        "  :category_id, :item_code, :name, :description, :item_type, "
                        "  :primary_uom_id, :secondary_uom_id, :hsn_code, :sku, :barcode_type, "
                        "  :has_batch, :has_serial, :has_expiry, :shelf_life_days, "
                        "  :safety_stock, :reorder_level, :reorder_qty, :lead_time_days, "
                        "  :min_order_qty, :max_order_qty, :weight, :weight_uom, :volume, :volume_uom, "
                        "  :purchase_price, :selling_price, :mrp, :tax_rate, :cgst, :sgst, :igst, "
                        "  :brand, :manufacturer, 1, NOW(), NOW()"
                        ")"
                    ),
                    {
                        "category_id": category_id,
                        "item_code": item_code,
                        "name": item_name,
                        "description": description,
                        "item_type": item_type,
                        "primary_uom_id": primary_uom_id,
                        "secondary_uom_id": secondary_uom_id,
                        "hsn_code": hsn_code,
                        "sku": sku,
                        "barcode_type": barcode_type,
                        "has_batch": has_batch,
                        "has_serial": has_serial,
                        "has_expiry": has_expiry,
                        "shelf_life_days": shelf_life_days,
                        "safety_stock": float(safety_stock),
                        "reorder_level": float(reorder_level),
                        "reorder_qty": float(reorder_qty),
                        "lead_time_days": lead_time_days,
                        "min_order_qty": float(min_order_qty),
                        "max_order_qty": float(max_order_qty),
                        "weight": float(weight) if weight else None,
                        "weight_uom": weight_uom,
                        "volume": float(volume) if volume else None,
                        "volume_uom": volume_uom,
                        "purchase_price": float(purchase_price),
                        "selling_price": float(selling_price),
                        "mrp": float(mrp),
                        "tax_rate": float(tax_rate),
                        "cgst": float(cgst),
                        "sgst": float(sgst),
                        "igst": float(igst),
                        "brand": brand,
                        "manufacturer": manufacturer,
                    },
                )
                iid = conn.execute(text("SELECT LAST_INSERT_ID()")).scalar()
                item_code_to_id[item_code.upper()] = iid
                item_count += 1

                if item_count % 50 == 0:
                    print(f"    ... {item_count} items imported so far")

                # Collect packing data
                packing_name = clean_str(row.get("Packing Name"))
                packing_uom_text = clean_str(row.get("Packing UOM"))
                qty_per_pack = parse_decimal(row.get("Qty Per Pack"))
                if packing_name and qty_per_pack > 0:
                    packing_data.append({
                        "item_code": item_code,
                        "packing_name": packing_name,
                        "packing_uom_text": packing_uom_text,
                        "qty_per_pack": qty_per_pack,
                    })

        print(f"\n    Imported {item_count} items ({skipped} duplicates skipped).")

        # ==================================================================
        # STEP 2f -- Item Packing
        # ==================================================================
        print(f"\n[2f] Importing Item Packing ({len(packing_data)} entries) ...")

        pack_count = 0
        for p in packing_data:
            iid = item_code_to_id.get(p["item_code"].upper())
            if not iid:
                continue

            packing_uom_id = resolve_uom(p["packing_uom_text"])
            if not packing_uom_id:
                # create fallback UOM for packing
                fallback_name = p["packing_uom_text"] or "Pack"
                fallback_abbr = (p["packing_uom_text"] or "Pk")[:10]
                print(f"    WARN: Packing UOM '{p['packing_uom_text']}' not found, creating '{fallback_name}'")
                conn.execute(
                    text(
                        "INSERT INTO uom (name, abbreviation, is_active, created_at) "
                        "VALUES (:name, :abbr, 1, NOW())"
                    ),
                    {"name": fallback_name, "abbr": fallback_abbr},
                )
                packing_uom_id = conn.execute(text("SELECT LAST_INSERT_ID()")).scalar()
                uom_name_to_id[fallback_name.lower()] = packing_uom_id
                uom_abbr_to_id[fallback_abbr.lower()] = packing_uom_id

            conn.execute(
                text(
                    "INSERT INTO item_packing (item_id, packing_name, packing_uom_id, qty_per_pack, is_default) "
                    "VALUES (:item_id, :name, :uom_id, :qty, 0)"
                ),
                {
                    "item_id": iid,
                    "name": p["packing_name"],
                    "uom_id": packing_uom_id,
                    "qty": float(p["qty_per_pack"]),
                },
            )
            pack_count += 1

        print(f"    Imported {pack_count} item packing entries.")

        # ==================================================================
        # STEP 2g -- Vendors
        # ==================================================================
        print("\n[2g] Importing Vendors ...")

        vendor_sheet = read_sheet(wb, "Vendors")
        vendor_count = 0

        for row in vendor_sheet:
            vendor_code = clean_str(row.get("Vendor Code"))
            vendor_name = clean_str(row.get("Vendor Name"))
            if not vendor_code or not vendor_name:
                continue

            vendor_type_raw = clean_str(row.get("Vendor Type"))
            vendor_type = "material"
            if vendor_type_raw:
                vt = vendor_type_raw.strip().lower()
                if vt in ("material", "transport", "service", "both"):
                    vendor_type = vt

            is_transport = parse_bool(row.get("Is Transport Vendor"))

            conn.execute(
                text(
                    "INSERT INTO vendors ("
                    "  vendor_code, name, contact_person, email, phone, alt_phone, "
                    "  address_line1, address_line2, city, state, pincode, country, "
                    "  gst_number, pan_number, bank_name, bank_account, bank_ifsc, "
                    "  payment_terms_days, credit_limit, vendor_type, is_transport_vendor, "
                    "  is_active, created_at, updated_at"
                    ") VALUES ("
                    "  :vendor_code, :name, :contact_person, :email, :phone, :alt_phone, "
                    "  :addr1, :addr2, :city, :state, :pincode, :country, "
                    "  :gst, :pan, :bank_name, :bank_account, :bank_ifsc, "
                    "  :payment_terms, :credit_limit, :vendor_type, :is_transport, "
                    "  1, NOW(), NOW()"
                    ")"
                ),
                {
                    "vendor_code": vendor_code,
                    "name": vendor_name,
                    "contact_person": clean_str(row.get("Contact Person")),
                    "email": clean_str(row.get("Email")),
                    "phone": clean_str(row.get("Phone")),
                    "alt_phone": clean_str(row.get("Alt Phone")),
                    "addr1": clean_str(row.get("Address Line 1")),
                    "addr2": clean_str(row.get("Address Line 2")),
                    "city": clean_str(row.get("City")),
                    "state": clean_str(row.get("State")),
                    "pincode": clean_str(row.get("Pincode")),
                    "country": clean_str(row.get("Country")) or "India",
                    "gst": clean_str(row.get("GST Number")),
                    "pan": clean_str(row.get("PAN Number")),
                    "bank_name": clean_str(row.get("Bank Name")),
                    "bank_account": clean_str(row.get("Bank Account")),
                    "bank_ifsc": clean_str(row.get("Bank IFSC")),
                    "payment_terms": parse_int(row.get("Payment Terms Days"), 30),
                    "credit_limit": float(parse_decimal(row.get("Credit Limit"))),
                    "vendor_type": vendor_type,
                    "is_transport": is_transport,
                },
            )
            vendor_count += 1
            print(f"    Vendor: {vendor_code} - {vendor_name}")

        print(f"    Imported {vendor_count} vendors.")

    # ==================================================================
    # Summary
    # ==================================================================
    print("\n" + "=" * 60)
    print("IMPORT COMPLETE")
    print("=" * 60)
    print(f"  UOMs:            {len(uom_name_to_id)}")
    print(f"  UOM Conversions: {conv_count}")
    print(f"  Projects:        {proj_count}")
    print(f"  Categories:      {len(cat_code_to_id)}")
    print(f"  Items:           {item_count} ({skipped} dups skipped)")
    print(f"  Item Packing:    {pack_count}")
    print(f"  Vendors:         {vendor_count}")
    print("=" * 60)


if __name__ == "__main__":
    main()
